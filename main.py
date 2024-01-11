
from openpyxl import Workbook, load_workbook
import os
import re
import datetime


DB = "employee_assets.xlsx"

def add_employ():
    emp_name = input("please enter your name:  ")
    
    emp_id = str(input("please enter your id: "))

    emp_id = id_validation(emp_id)
    exist = search_by_emp_id(2, emp_id)
    if exist == None:
        print("New Employee has no prior records, Please add your assets")
        result = add_asset(emp_id)
        save_data([emp_name] + result)

def add_asset(emp_id = "0"):
    
    if id_validation(emp_id) == emp_id: 
        asset = input("Asset type: \n1 - Laptop\n2 - HeadPhone\n3 - Mobile\n4 - Monitor\n")
        asset_sn = input("please enter asset serial number: ")
        note = input("note: ")
        timestamp = datetime.datetime.now()
        return [emp_id, asset, asset_sn, note, timestamp]
    
def open_workbook(file_name = "employee_assets.xlsx"):
    script_dir = os.path.dirname(os.path.abspath(__file__))  # Get the directory of the script
    file_path = os.path.join(script_dir, file_name)  # Combine with the relative file path
    try:
        wb = load_workbook(file_path)
        return wb
    except FileNotFoundError:
        print(f"Error while opening: File '{file_path}' not found.")
        return None
    except Exception as e:
        print(f"Error while opening: {e}")
        return None
    
def append_to_workbook(wb, data_to_append):
    try:
        ws = wb.active
        ws.append(data_to_append)  # Append a row with the provided data
        wb.save(DB)  # Save the changes
        print("Row appended and workbook closed successfully.")
        return True
    except Exception as e:
        print(f"Error while appending: {e}")
        return False
    
def close_workbook(wb):
    try:
        ws = wb.active
        wb.close()  # Close the workbook
        return True
    except Exception as e:
        print(f"Error while closing: {e}")
        return False
    
def save_data(data):
    confirm = input("Are you sure you want to append the following record ?\n"+str(data)+"\n(y/n)")
    while(confirm not in ["y","n"]):
        print("wrong choice please enter \'y\' or \'n\'")
        confirm = input("Are you sure you want to append the following record ?\n"+str(data)+"\n(y/n)")
    if confirm == "y":
            wb = open_workbook()
            result = append_to_workbook(wb, data)
            close_workbook(wb)
            return result
    elif confirm == "n":
        return False
    
def validate_14_digit(input_string):
    pattern = re.compile(r'^\d{14}$')
    match = pattern.match(input_string) # If there is a match, the input is a 14-digit number
    return bool(match)

def id_validation(emp_id):
    while( not validate_14_digit(emp_id)):
        emp_id = str(input("please enter a valid id, should be 14 digit number: "))
    return emp_id

def search_by_emp_id(target_emp_id):
    try:
        wb = open_workbook()
        ws = wb["EmployeeAssets"]
        id_col = ws.iter_rows(min_row= 2, min_col= 1, max_col= 6)
        matched = []
        for row in id_col:
            for cell in row: 
                if cell.value == target_emp_id:
                    matched.append([cell.value for cell in row])
        return matched if matched else None
    except Exception as e:
        print(f"Error: {e}")
        return None
 
def user():
    #---------------------------------------------
    # user enter 14 digit id ..if new id we register name
    #---------------------------------------------
    emp_id = str(input("please enter your id : "))
    emp_id = id_validation(emp_id)
    exist = search_by_emp_id( emp_id)
    if exist == None:
        print("New Employee has no prior records, Please add your Name")
        emp_name = input("please enter your name:  ")
        exist = []
        exist.append([emp_id, emp_name, "N/A", "N/A", "N/A", datetime.datetime.now() ]) 
    print("Hello, "+exist[0][1]+"!")
    print("you have "+str(len(exist) - 1)+" assets")
    if len(exist) :
        print("Your assets are :")
        print("Type","S/N","Note","Time stamp",sep="\t")
        for i in exist[:-1]:
            print(i[2],i[3],i[4],i[5],sep="\t")
    
def main_menu():
    wb = open_workbook()
    print("1 - enter user data")
    print("2 - list by:")
    x = int(input('Enter a number: '))
    
    if x == 1:
        user()
        append_to_workbook(wb, data_to_append)


if __name__ == "__main__":
    # x = "45678901254567"
    # print(search_by_emp_id(x))

    main_menu()
    # data = ["John Doe", "123456","Laptop", "1000", "2023-01-01"]
    # print(save_data(data))